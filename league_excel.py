from riot_tracker import RiotTracker
from player_stats import PlayerStats
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
import os

def parse_match(match_dict, puuid):
    #finding what index the player is in the participant array
 
    puuid_list = match_dict["metadata"]["participants"]

    for x in range (0, 10):
        if (puuid_list[x] == puuid):
            idx = x
            break
    
    player_dict = match_dict["info"]["participants"][idx]

    #parsing role

    if player_dict["teamPosition"] == "UTILITY":
        role = "SUPPORT"
    else:
        role = player_dict["teamPosition"]

    champion = player_dict["championName"]
    victory = player_dict["win"]
    kills = player_dict["kills"]
    deaths = player_dict["deaths"]
    assists = player_dict["assists"]
    kda = (kills + assists) / deaths

    #parsing time

    time_milliseconds = match_dict["info"]["gameDuration"] #time in MILLISECONDS MUST CONVERT

    time_minutes = int((time_milliseconds / (1000 * 60) ) % 60)
    time_seconds = int(time_milliseconds / 1000) % 60

    time_string = f"{time_minutes}:{time_seconds}"

    if len(str(time_seconds)) == 1:
        time_string = time_string[:4] + "0" + time_string[4:]

    cs_total = player_dict["totalMinionsKilled"] + player_dict["neutralMinionsKilled"]
    cs_per_min = round(cs_total / (time_milliseconds / 60000), 1) #rounded to one decimal place
    champ_physical_damage = player_dict["physicalDamageDealtToChampions"]
    champ_magic_damage = player_dict["magicDamageDealtToChampions"]
    champ_true_damage = player_dict["trueDamageDealtToChampions"]
    champ_total_damage = player_dict["totalDamageDealtToChampions"]
    gold = player_dict["goldEarned"]

    #creating PlayerStats object for easy management of spreadsheet

    player_stats = PlayerStats(role, champion, victory, kills, deaths, assists, kda, time_milliseconds, time_string, cs_total, cs_per_min, champ_physical_damage, champ_magic_damage, champ_true_damage, champ_total_damage, gold)

    return player_stats

def create_layout(sheet):
    sheet.title = "Stats Page"
    sheet["A1"] = "Role"
    sheet["B1"] = "Champion"
    sheet["D1"] = "Kills"
    sheet["E1"] = "Deaths"
    sheet["F1"] = "Assists"
    sheet["G1"] = "KDA"
    sheet["H1"] = "Result"
    sheet["I1"] = "Time"
    sheet["J1"] = "Total CS"
    sheet["K1"] = "CS/m"
    sheet["L1"] = "Gold"
    sheet["N1"] = "Physical Damage"
    sheet["P1"] = "Magic Damage"
    sheet["R1"] = "True Damage"
    sheet["T1"] = "Total Damage"
    
    return sheet;

def add_stats(sheet, player_stats: PlayerStats, row_idx: int):

    # COLOR CODES
    # GOOD = 96DE91, NEUTRAL = EBF5A7, BAD = F5A7A7

    good = PatternFill("solid", fgColor="96DE91")
    neutral = PatternFill("solid", fgColor="EBF5A7")
    bad = PatternFill("solid", fgColor="F5A7A7")

    sheet.cell(row=row_idx, column=1).value = player_stats.role
    sheet.cell(row=row_idx, column=2).value = player_stats.champion
    sheet.cell(row=row_idx, column=4).value = player_stats.kills
    sheet.cell(row=row_idx, column=5).value = player_stats.deaths
    sheet.cell(row=row_idx, column=6).value = player_stats.assists
    sheet.cell(row=row_idx, column=7).value = player_stats.kda

    if player_stats.kda < 2:
        sheet.cell(row=row_idx, column=7).fill = bad
    elif player_stats.kda >= 4: 
        sheet.cell(row=row_idx, column=7).fill = good
    else:
        sheet.cell(row=row_idx, column=7).fill = neutral
    

    sheet.cell(row=row_idx, column=8).value = ("Defeat", "Victory")[player_stats.victory]

    if player_stats.victory:
        sheet.cell(row=row_idx, column=8).fill = good
    else:
        sheet.cell(row=row_idx, column=8).fill = bad

    sheet.cell(row=row_idx, column=9).value = player_stats.time_string
    sheet.cell(row=row_idx, column=10).value = player_stats.cs_total
    sheet.cell(row=row_idx, column=11).value = player_stats.cs_per_min

    if player_stats.cs_per_min < 6:
        sheet.cell(row=row_idx, column=11).fill = bad
    elif player_stats.cs_per_min >= 8:
        sheet.cell(row=row_idx, column=11).fill = good
    else:
        sheet.cell(row=row_idx, column=11).fill = neutral

    sheet.cell(row=row_idx, column=12).value = player_stats.gold
    sheet.cell(row=row_idx, column=14).value = player_stats.champ_physical_damage
    sheet.cell(row=row_idx, column=16).value = player_stats.champ_magic_damage
    sheet.cell(row=row_idx, column=18).value = player_stats.champ_true_damage
    sheet.cell(row=row_idx, column=20).value = player_stats.champ_total_daamge

    
    
def create_spreadsheet(riot_tracker: RiotTracker, match_list, puuid):
    match_list.reverse()

    workbook = Workbook()
    sheet = workbook.active

    sheet = create_layout(sheet)

    row_idx = 2

    for match_id in match_list:
        match_dict = riot_tracker.get_match(match_id)
        player_stats = parse_match(match_dict, puuid)
        add_stats(sheet, player_stats, row_idx)
        row_idx += 1
    

    workbook.save(filename="test.xlsx")




def update_spreadsheet():
    pass

def main():
    api_key = input("Enter Riot API Key (refer to *github link* if you don't know how to get this): ")
    name = input("Enter summoner name: ")
    file_name = input("Enter file name (if you don't have an existing one, we will automatically create one with this name): ")

    riot_tracker = RiotTracker(api_key)
    puuid = riot_tracker.get_puuid(name)
    match_list = riot_tracker.get_match_history(puuid, 10)

    file_list = os.listdir(".")

    fileFound = False

    for file in file_list:
        if file == file_name:
            fileFound = True

    if (fileFound == False):
        create_spreadsheet(riot_tracker, match_list, puuid)
    else:
        update_spreadsheet()
    


if __name__ == "__main__":
    main()
